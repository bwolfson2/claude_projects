#!/usr/bin/env python3
"""
Sync JSON CRM data to Excel spreadsheet (JSON → Excel)

This script reads deals.json and projects.json, then updates the VFT-Master-Tracker.xlsx
spreadsheet while preserving all formatting and structure.

Usage:
    python sync_to_xlsx.py [--verbose]

Dependencies:
    openpyxl >= 3.0
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
import argparse

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    exit(1)


class SyncLog:
    """Manages audit logging for sync operations."""

    def __init__(self, log_path: str):
        self.log_path = log_path
        self.entries = []

    def add(self, message: str, level: str = "INFO"):
        """Add a log entry."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        entry = f"[{timestamp}] {level}: {message}"
        self.entries.append(entry)
        if level == "ERROR":
            print(f"  ❌ {message}")
        elif level == "WARNING":
            print(f"  ⚠️  {message}")
        else:
            print(f"  ✓ {message}")

    def save(self):
        """Save log to file (append mode)."""
        os.makedirs(os.path.dirname(self.log_path), exist_ok=True)
        with open(self.log_path, 'a') as f:
            for entry in self.entries:
                f.write(entry + "\n")
        print(f"\n📝 Sync log written to {self.log_path}")


class ExcelSyncer:
    """Synchronizes JSON data to Excel spreadsheet."""

    def __init__(self, excel_path: str, log: SyncLog, verbose: bool = False):
        self.excel_path = excel_path
        self.log = log
        self.verbose = verbose
        self.changes_count = {"added": 0, "updated": 0, "cleared": 0}

    def load_workbook(self):
        """Load the Excel workbook."""
        if not os.path.exists(self.excel_path):
            self.log.add(f"Excel file not found: {self.excel_path}", "ERROR")
            raise FileNotFoundError(self.excel_path)

        try:
            self.wb = load_workbook(self.excel_path)
            self.log.add(f"Loaded Excel workbook: {self.excel_path}")
            return True
        except Exception as e:
            self.log.add(f"Failed to load Excel: {str(e)}", "ERROR")
            raise

    def load_json(self, json_path: str) -> Dict[str, Any]:
        """Load JSON file."""
        if not os.path.exists(json_path):
            self.log.add(f"JSON file not found: {json_path} (will skip)", "WARNING")
            return None

        try:
            with open(json_path, 'r') as f:
                data = json.load(f)
            self.log.add(f"Loaded JSON file: {json_path}")
            return data
        except Exception as e:
            self.log.add(f"Failed to load JSON {json_path}: {str(e)}", "ERROR")
            raise

    def sync_dd_pipeline(self, deals_data: Dict[str, Any]):
        """Sync deals data to DD Pipeline tab."""
        print("\n📊 Syncing DD Pipeline tab...")

        if not deals_data or "companies" not in deals_data:
            self.log.add("No companies found in deals.json", "WARNING")
            return

        ws = self.wb["DD Pipeline"]
        companies = deals_data.get("companies", [])

        self.log.add(f"Found {len(companies)} companies to sync")

        # Field mapping: (json_field_path, excel_column)
        # Use dot notation for nested fields, e.g., "diligence.commercial"
        field_mapping = [
            ("company_name", "A"),
            ("stage", "B"),
            ("status", "C"),
            ("decision_posture", "D"),
            ("sector", "E"),
            ("round", "F"),
            ("raise_usd", "G"),
            ("valuation_cap_usd", "H"),
            ("owner", "I"),
            ("priority", "J"),
            ("last_touch", "K"),
            ("next_action", "L"),
            ("next_action_owner", "M"),
            ("next_action_due", "N"),
            ("thesis", "O"),
            ("diligence.commercial", "P"),
            ("diligence.product_technical", "Q"),
            ("diligence.finance_legal", "R"),
            ("diligence.memo", "S"),
            ("open_questions", "T"),
            ("assumptions", "U"),
        ]

        # Clear existing data rows (keep headers)
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).value = None

        # Write companies
        for row_idx, company in enumerate(companies, start=2):
            for json_field, col_letter in field_mapping:
                value = self._get_nested_value(company, json_field)

                # Format the value for Excel
                if json_field.startswith("diligence."):
                    # Status fields - keep as-is
                    cell_value = value or ""
                elif json_field == "open_questions" or json_field == "assumptions":
                    # Join arrays with newlines
                    if isinstance(value, list):
                        cell_value = "\n".join(value) if value else ""
                    else:
                        cell_value = value or ""
                else:
                    # Regular field
                    cell_value = value or ""

                # Write to cell
                cell = ws[f"{col_letter}{row_idx}"]
                cell.value = cell_value

                # Set alignment for multi-line cells
                if isinstance(cell_value, str) and "\n" in cell_value:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            self.changes_count["added"] += 1

        self.log.add(f"Synced {len(companies)} companies to DD Pipeline tab")

    def sync_project_management(self, projects_data: Optional[Dict[str, Any]]):
        """Sync projects data to Project Management tab."""
        print("\n📋 Syncing Project Management tab...")

        ws = self.wb["Project Management"]

        if not projects_data:
            self.log.add("No projects.json found (Project Management tab unchanged)", "WARNING")
            return

        if isinstance(projects_data, list):
            projects = projects_data
        elif isinstance(projects_data, dict) and "projects" in projects_data:
            projects = projects_data["projects"]
        else:
            self.log.add("Unexpected projects.json structure", "WARNING")
            return

        self.log.add(f"Found {len(projects)} projects to sync")

        # Field mapping for projects
        field_mapping = [
            ("name", "A"),
            ("category", "B"),
            ("status", "C"),
            ("priority", "D"),
            ("owner", "E"),
            ("start_date", "F"),
            ("target_date", "G"),
            ("description", "H"),
            ("next_action", "I"),
            ("next_action_owner", "J"),
            ("next_action_due", "K"),
            ("artifacts", "L"),
            ("notes", "M"),
        ]

        # Clear existing data rows
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).value = None

        # Write projects
        for row_idx, project in enumerate(projects, start=2):
            for json_field, col_letter in field_mapping:
                value = self._get_nested_value(project, json_field)

                # Format the value
                if json_field == "artifacts":
                    # Serialize artifacts
                    if isinstance(value, dict):
                        cell_value = "\n".join([f"{k}: {v}" for k, v in value.items()]) if value else ""
                    elif isinstance(value, list):
                        cell_value = "\n".join(value) if value else ""
                    else:
                        cell_value = value or ""
                elif isinstance(value, list):
                    cell_value = "\n".join(value) if value else ""
                else:
                    cell_value = value or ""

                # Write to cell
                cell = ws[f"{col_letter}{row_idx}"]
                cell.value = cell_value

                # Set alignment for multi-line cells
                if isinstance(cell_value, str) and "\n" in cell_value:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            self.changes_count["added"] += 1

        self.log.add(f"Synced {len(projects)} projects to Project Management tab")

    def _get_nested_value(self, obj: Dict[str, Any], path: str) -> Any:
        """Get value from nested dictionary using dot notation."""
        parts = path.split(".")
        current = obj

        for part in parts:
            if isinstance(current, dict):
                current = current.get(part)
            else:
                return None

        return current

    def save_workbook(self):
        """Save the workbook."""
        try:
            self.wb.save(self.excel_path)
            self.log.add(f"Saved Excel workbook: {self.excel_path}")
            return True
        except Exception as e:
            self.log.add(f"Failed to save Excel: {str(e)}", "ERROR")
            raise

    def print_summary(self):
        """Print sync summary."""
        print("\n" + "=" * 60)
        print("SYNC SUMMARY")
        print("=" * 60)
        print(f"Companies synced:  {self.changes_count['added']}")
        print(f"Excel file:        {self.excel_path}")
        print("=" * 60)


def main():
    """Main sync function."""
    parser = argparse.ArgumentParser(
        description="Sync JSON CRM data to Excel spreadsheet"
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Print verbose output"
    )

    args = parser.parse_args()

    # Determine file paths
    script_dir = Path(__file__).parent.parent
    repo_root = script_dir.parent.parent

    deals_json = repo_root / "fund" / "crm" / "deals.json"
    projects_json = repo_root / "projects" / "projects.json"
    excel_file = repo_root / "fund" / "VFT-Master-Tracker.xlsx"
    log_file = script_dir / "references" / "sync-log.txt"

    print("🔄 Tracker Sync: JSON → Excel")
    print("=" * 60)
    print(f"Deals JSON:      {deals_json}")
    print(f"Projects JSON:   {projects_json}")
    print(f"Excel file:      {excel_file}")
    print("=" * 60)

    # Initialize sync log
    log = SyncLog(str(log_file))
    log.add("=== Starting JSON → Excel sync ===")

    try:
        # Create syncer
        syncer = ExcelSyncer(str(excel_file), log, args.verbose)

        # Load workbook
        syncer.load_workbook()

        # Load JSON files
        deals_data = syncer.load_json(str(deals_json))
        projects_data = syncer.load_json(str(projects_json))

        # Perform syncs
        if deals_data:
            syncer.sync_dd_pipeline(deals_data)

        if projects_data:
            syncer.sync_project_management(projects_data)

        # Save workbook
        syncer.save_workbook()

        # Print summary
        syncer.print_summary()

        # Save log
        log.add("=== Sync completed successfully ===")
        log.save()

        print("\n✅ Sync completed successfully!")

    except Exception as e:
        log.add(f"Sync failed: {str(e)}", "ERROR")
        log.save()
        print(f"\n❌ Sync failed: {str(e)}")
        exit(1)


if __name__ == "__main__":
    main()
