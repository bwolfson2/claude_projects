#!/usr/bin/env python3
"""
Sync Excel spreadsheet data back to JSON files (Excel → JSON)

This script reads the VFT-Master-Tracker.xlsx spreadsheet and syncs changes back to
deals.json and projects.json. Useful when the spreadsheet is shared for collaborative
editing in Google Sheets.

CAUTION: This is a reverse sync that modifies JSON files based on Excel data.
Always review changes with --dry-run first.

Usage:
    python sync_from_xlsx.py --dry-run       # Preview changes without writing
    python sync_from_xlsx.py                 # Apply changes
    python sync_from_xlsx.py --verbose       # Verbose output

Dependencies:
    openpyxl >= 3.0
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import argparse
import copy

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    exit(1)


class SyncLog:
    """Manages audit logging for sync operations."""

    def __init__(self, log_path: str):
        self.log_path = log_path
        self.entries = []

    def add(self, message: str, level: str = "INFO"):
        """Add a log entry."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        entry = f"[{timestamp}] {level}: {message}"
        self.entries.append(entry)
        if level == "ERROR":
            print(f"  ❌ {message}")
        elif level == "WARNING":
            print(f"  ⚠️  {message}")
        else:
            print(f"  ✓ {message}")

    def save(self):
        """Save log to file (append mode)."""
        os.makedirs(os.path.dirname(self.log_path), exist_ok=True)
        with open(self.log_path, 'a') as f:
            for entry in self.entries:
                f.write(entry + "\n")


class JSONSyncer:
    """Synchronizes Excel data back to JSON files."""

    def __init__(self, excel_path: str, log: SyncLog, dry_run: bool = False, verbose: bool = False):
        self.excel_path = excel_path
        self.log = log
        self.dry_run = dry_run
        self.verbose = verbose
        self.changes_count = {"deals_updated": 0, "projects_updated": 0}

    def load_workbook(self):
        """Load the Excel workbook."""
        if not os.path.exists(self.excel_path):
            self.log.add(f"Excel file not found: {self.excel_path}", "ERROR")
            raise FileNotFoundError(self.excel_path)

        try:
            self.wb = load_workbook(self.excel_path)
            self.log.add(f"Loaded Excel workbook: {self.excel_path}")
            return True
        except Exception as e:
            self.log.add(f"Failed to load Excel: {str(e)}", "ERROR")
            raise

    def load_json(self, json_path: str) -> Optional[Dict[str, Any]]:
        """Load JSON file."""
        if not os.path.exists(json_path):
            self.log.add(f"JSON file not found: {json_path}", "WARNING")
            return None

        try:
            with open(json_path, 'r') as f:
                data = json.load(f)
            self.log.add(f"Loaded JSON: {json_path}")
            return data
        except Exception as e:
            self.log.add(f"Failed to load JSON {json_path}: {str(e)}", "ERROR")
            raise

    def sync_deals_from_excel(self, deals_path: str):
        """Sync DD Pipeline tab back to deals.json."""
        print("\n📊 Syncing DD Pipeline tab → deals.json...")

        # Load current deals
        deals_data = self.load_json(deals_path)
        if not deals_data:
            self.log.add("No existing deals.json to update", "WARNING")
            return

        # Create index of existing companies by company_name
        company_index = {c["company_name"]: c for c in deals_data.get("companies", [])}
        original_data = copy.deepcopy(deals_data)

        ws = self.wb["DD Pipeline"]

        # Field mapping
        field_mapping = [
            ("company_name", "A"),
            ("stage", "B"),
            ("status", "C"),
            ("decision_posture", "D"),
            ("sector", "E"),
            ("round", "F"),
            ("raise_usd", "G"),
            ("valuation_cap_usd", "H"),
            ("owner", "I"),
            ("priority", "J"),
            ("last_touch", "K"),
            ("next_action", "L"),
            ("next_action_owner", "M"),
            ("next_action_due", "N"),
            ("thesis", "O"),
            ("diligence.commercial", "P"),
            ("diligence.product_technical", "Q"),
            ("diligence.finance_legal", "R"),
            ("diligence.memo", "S"),
            ("open_questions", "T"),
            ("assumptions", "U"),
        ]

        # Iterate through data rows
        for row_idx in range(2, ws.max_row + 1):
            company_name = ws[f"A{row_idx}"].value

            if not company_name:
                continue

            # Find or create company record
            if company_name in company_index:
                company = company_index[company_name]
            else:
                # Create new company
                company = {"company_name": company_name}
                if "companies" not in deals_data:
                    deals_data["companies"] = []
                deals_data["companies"].append(company)
                company_index[company_name] = company
                self.log.add(f"Creating new company: {company_name}")

            # Update fields
            for json_field, col_letter in field_mapping:
                cell_value = ws[f"{col_letter}{row_idx}"].value

                # Skip if empty
                if cell_value is None or cell_value == "":
                    continue

                # Parse and set value
                if json_field == "open_questions" or json_field == "assumptions":
                    # Split multi-line cells
                    if isinstance(cell_value, str):
                        value = [line.strip() for line in cell_value.split("\n") if line.strip()]
                    else:
                        value = cell_value
                elif json_field.startswith("diligence."):
                    # Handle nested diligence fields
                    if "diligence" not in company:
                        company["diligence"] = {}
                    field_name = json_field.split(".")[1]
                    company["diligence"][field_name] = cell_value
                    continue
                else:
                    # Try to convert to appropriate type
                    value = self._parse_value(json_field, cell_value)

                self._set_nested_value(company, json_field, value)

            if self.verbose:
                self.log.add(f"Updated: {company_name}")

            self.changes_count["deals_updated"] += 1

        # Check if data changed
        if deals_data != original_data:
            self.log.add(f"Changes detected in {self.changes_count['deals_updated']} companies")
            if not self.dry_run:
                self._save_json(deals_path, deals_data)
                self.log.add(f"Saved updated deals.json")
        else:
            self.log.add("No changes detected in deals.json")

    def sync_projects_from_excel(self, projects_path: str):
        """Sync Project Management tab back to projects.json."""
        print("\n📋 Syncing Project Management tab → projects.json...")

        ws = self.wb["Project Management"]

        # Check if there's any data
        has_data = False
        for row_idx in range(2, ws.max_row + 1):
            if ws[f"A{row_idx}"].value:
                has_data = True
                break

        if not has_data:
            self.log.add("No project data found in spreadsheet")
            return

        # Load or create projects
        projects_data = self.load_json(projects_path)
        if not projects_data:
            projects_data = {"projects": []}

        original_data = copy.deepcopy(projects_data)

        # Ensure projects is a list
        if isinstance(projects_data, list):
            projects = projects_data
        else:
            projects = projects_data.get("projects", [])

        # Create index by name
        project_index = {p.get("name"): p for p in projects if "name" in p}

        # Field mapping
        field_mapping = [
            ("name", "A"),
            ("category", "B"),
            ("status", "C"),
            ("priority", "D"),
            ("owner", "E"),
            ("start_date", "F"),
            ("target_date", "G"),
            ("description", "H"),
            ("next_action", "I"),
            ("next_action_owner", "J"),
            ("next_action_due", "K"),
            ("artifacts", "L"),
            ("notes", "M"),
        ]

        # Iterate through data rows
        for row_idx in range(2, ws.max_row + 1):
            project_name = ws[f"A{row_idx}"].value

            if not project_name:
                continue

            # Find or create project
            if project_name in project_index:
                project = project_index[project_name]
            else:
                project = {"name": project_name}
                projects.append(project)
                project_index[project_name] = project
                self.log.add(f"Creating new project: {project_name}")

            # Update fields
            for json_field, col_letter in field_mapping:
                cell_value = ws[f"{col_letter}{row_idx}"].value

                if cell_value is None or cell_value == "":
                    continue

                # Parse value
                if json_field == "artifacts":
                    # Parse "name: url" format
                    if isinstance(cell_value, str):
                        artifacts = {}
                        for line in cell_value.split("\n"):
                            if ":" in line:
                                key, val = line.split(":", 1)
                                artifacts[key.strip()] = val.strip()
                        value = artifacts if artifacts else cell_value
                    else:
                        value = cell_value
                else:
                    value = self._parse_value(json_field, cell_value)

                project[json_field] = value

            if self.verbose:
                self.log.add(f"Updated: {project_name}")

            self.changes_count["projects_updated"] += 1

        # Save if changed
        if isinstance(projects_data, list):
            updated_data = projects
        else:
            projects_data["projects"] = projects
            updated_data = projects_data

        if updated_data != original_data:
            self.log.add(f"Changes detected in {self.changes_count['projects_updated']} projects")
            if not self.dry_run:
                self._save_json(projects_path, updated_data)
                self.log.add(f"Saved updated projects.json")
        else:
            self.log.add("No changes detected in projects.json")

    def _parse_value(self, field_name: str, value: Any) -> Any:
        """Parse cell value to appropriate Python type."""
        if value is None or value == "":
            return None

        # Numeric fields
        if field_name in ["raise_usd", "valuation_cap_usd"]:
            try:
                return int(float(value))
            except (ValueError, TypeError):
                return value

        # Date fields - keep as string (YYYY-MM-DD)
        if field_name in ["last_touch", "next_action_due", "start_date", "target_date"]:
            return str(value) if value else None

        # String fields
        return str(value) if value else None

    def _set_nested_value(self, obj: Dict[str, Any], path: str, value: Any):
        """Set value in nested dictionary using dot notation."""
        parts = path.split(".")

        for part in parts[:-1]:
            if part not in obj:
                obj[part] = {}
            obj = obj[part]

        obj[parts[-1]] = value

    def _save_json(self, json_path: str, data: Any):
        """Save JSON file."""
        try:
            with open(json_path, 'w') as f:
                json.dump(data, f, indent=2)
            self.log.add(f"Saved: {json_path}")
        except Exception as e:
            self.log.add(f"Failed to save {json_path}: {str(e)}", "ERROR")
            raise

    def print_summary(self):
        """Print sync summary."""
        print("\n" + "=" * 60)
        if self.dry_run:
            print("DRY RUN - PREVIEW (NO CHANGES WRITTEN)")
        else:
            print("SYNC SUMMARY")
        print("=" * 60)
        print(f"Deals updated:    {self.changes_count['deals_updated']}")
        print(f"Projects updated: {self.changes_count['projects_updated']}")
        print("=" * 60)


def main():
    """Main sync function."""
    parser = argparse.ArgumentParser(
        description="Sync Excel spreadsheet data back to JSON files"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview changes without writing"
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Print verbose output"
    )

    args = parser.parse_args()

    # Determine file paths
    script_dir = Path(__file__).parent.parent
    repo_root = script_dir.parent.parent

    excel_file = repo_root / "fund" / "VFT-Master-Tracker.xlsx"
    deals_json = repo_root / "fund" / "crm" / "deals.json"
    projects_json = repo_root / "projects" / "projects.json"
    log_file = script_dir / "references" / "sync-log.txt"

    print("🔄 Tracker Sync: Excel → JSON")
    print("=" * 60)
    print(f"Excel file:      {excel_file}")
    print(f"Deals JSON:      {deals_json}")
    print(f"Projects JSON:   {projects_json}")
    if args.dry_run:
        print("Mode:            DRY RUN (preview only)")
    print("=" * 60)

    # Initialize log
    log = SyncLog(str(log_file))
    mode = "DRY RUN" if args.dry_run else "WRITE"
    log.add(f"=== Starting Excel → JSON sync ({mode}) ===")

    try:
        # Create syncer
        syncer = JSONSyncer(str(excel_file), log, args.dry_run, args.verbose)

        # Load workbook
        syncer.load_workbook()

        # Perform syncs
        syncer.sync_deals_from_excel(str(deals_json))
        syncer.sync_projects_from_excel(str(projects_json))

        # Print summary
        syncer.print_summary()

        # Save log
        log.add(f"=== Sync completed ({mode}) ===")
        log.save()

        if args.dry_run:
            print("\n✅ Dry run completed. Review changes above.")
            print("   Run without --dry-run to apply changes.")
        else:
            print("\n✅ Sync completed successfully!")

    except Exception as e:
        log.add(f"Sync failed: {str(e)}", "ERROR")
        log.save()
        print(f"\n❌ Sync failed: {str(e)}")
        exit(1)


if __name__ == "__main__":
    main()
